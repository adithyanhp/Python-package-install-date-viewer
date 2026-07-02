"""
core/exporter.py

Exports a list of PackageInfo objects to CSV, Excel (.xlsx), or PDF.
Excel/PDF support is optional at runtime - if openpyxl or reportlab
aren't installed, those export methods raise a clear ImportError-based
exception rather than crashing the whole app.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from core.models import PackageInfo, PythonInterpreter
from utils.logger import get_logger

logger = get_logger(__name__)

_COLUMNS = ("Name", "Version", "Latest", "Size", "Location", "Summary", "Editable")


class ExportDependencyError(RuntimeError):
    """Raised when an optional export dependency is missing."""


def _rows(packages: list[PackageInfo]) -> list[tuple[str, str, str, str, str, str, str]]:
    rows = []
    for pkg in packages:
        rows.append(
            (
                pkg.name,
                pkg.version,
                pkg.latest_version or "",
                pkg.size_human,
                str(pkg.location) if pkg.location else "",
                pkg.summary,
                "Yes" if pkg.is_editable else "No",
            )
        )
    return rows


class Exporter:
    """Exports package lists to various file formats."""

    def export_csv(
        self, packages: list[PackageInfo], interpreter: PythonInterpreter, out_path: Path
    ) -> Path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow([f"Interpreter: {interpreter.display_name}"])
            writer.writerow([f"Exported: {datetime.now().isoformat(timespec='seconds')}"])
            writer.writerow([])
            writer.writerow(_COLUMNS)
            writer.writerows(_rows(packages))
        logger.info("Exported %d packages to CSV: %s", len(packages), out_path)
        return out_path

    def export_excel(
        self, packages: list[PackageInfo], interpreter: PythonInterpreter, out_path: Path
    ) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ExportDependencyError(
                "Excel export requires the 'openpyxl' package. Install it with: pip install openpyxl"
            ) from exc

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Packages"

        ws.append([f"Interpreter: {interpreter.display_name}"])
        ws.append([f"Exported: {datetime.now().isoformat(timespec='seconds')}"])
        ws.append([])
        header_row_idx = 4
        ws.append(list(_COLUMNS))

        header_fill = PatternFill(start_color="6C3FC5", end_color="6C3FC5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        for row in _rows(packages):
            ws.append(list(row))

        for col_idx, header in enumerate(_COLUMNS, start=1):
            max_len = max([len(header)] + [len(str(r[col_idx - 1])) for r in _rows(packages)] or [0])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(out_path)
        logger.info("Exported %d packages to Excel: %s", len(packages), out_path)
        return out_path

    def export_pdf(
        self, packages: list[PackageInfo], interpreter: PythonInterpreter, out_path: Path
    ) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:
            raise ExportDependencyError(
                "PDF export requires the 'reportlab' package. Install it with: pip install reportlab"
            ) from exc

        out_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(str(out_path), pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("PyPackage Manager Pro - Export Report", styles["Title"]),
            Paragraph(f"Interpreter: {interpreter.display_name}", styles["Normal"]),
            Paragraph(
                f"Exported: {datetime.now().isoformat(timespec='seconds')}", styles["Normal"]
            ),
            Spacer(1, 12),
        ]

        table_data = [list(_COLUMNS)] + [list(row) for row in _rows(packages)]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C3FC5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2ECFB")]),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        logger.info("Exported %d packages to PDF: %s", len(packages), out_path)
        return out_path
